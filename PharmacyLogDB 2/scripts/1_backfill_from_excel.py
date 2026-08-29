"""
1_backfill_from_excel.py — run this ONCE, before you start doing daily
imports. Loads every historical prescription from the current Excel
tracker's Daily Log tab, including whatever your team already filled
in for DOA / Insurance / Office / Demo / Notes.

Needs the 'openpyxl' add-on:  pip install openpyxl
"""

import sys

import config
import pharmacy_common as common

try:
    import openpyxl
except ImportError:
    print("Missing 'openpyxl'. Run this in Command Prompt, then try again:")
    print("    pip install openpyxl")
    sys.exit(1)


def main():
    if not config.EXCEL_FILE.exists():
        print(f"could not find the Excel file at {config.EXCEL_FILE}")
        print("Fix the EXCEL_FILE path in config.py.")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(config.EXCEL_FILE), data_only=True)
    if config.EXCEL_DAILY_LOG_SHEET not in wb.sheetnames:
        print(f"could not find a tab named '{config.EXCEL_DAILY_LOG_SHEET}'")
        print(f"Tabs found in this workbook: {', '.join(wb.sheetnames)}")
        print("Fix EXCEL_DAILY_LOG_SHEET in config.py to match exactly.")
        sys.exit(1)

    sheet = wb[config.EXCEL_DAILY_LOG_SHEET]
    rows, meta = common.read_daily_log_sheet(sheet)

    if rows is None:
        print(f"Could not find a table named '{config.EXCEL_TABLE_NAME}' on")
        print(f"'{config.EXCEL_DAILY_LOG_SHEET}', and couldn't recognize a plain")
        print("header row in the first 15 rows either. Here's what row 1 looks like:")
        row1 = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        print(f"  {row1}")
        print("Update EXCEL_TABLE_NAME or EXCEL_COLUMN_ALIASES in config.py and re-run.")
        sys.exit(1)

    if meta["used_fallback"]:
        print(f"Note: no table named '{config.EXCEL_TABLE_NAME}' found — fell back to")
        print("scanning for a plain header row instead.")

    if not rows:
        print("No prescription rows detected in the Daily Log tab.")
        print("Double-check EXCEL_DAILY_LOG_SHEET and EXCEL_TABLE_NAME in config.py.")
        sys.exit(1)

    conn = common.get_connection()
    common.ensure_schema(conn)
    new_count, filled_count = common.upsert_with_manual_fields(
        conn, rows, source_file=config.EXCEL_FILE.name
    )
    conn.close()

    print(f"Rows found: {len(rows)}  new: {new_count}  "
          f"(skipped {meta['skipped_separator']} 'Log Date:' separator rows)")
    if meta["unknown_date_count"]:
        print(f"Warning: {meta['unknown_date_count']} row(s) had no preceding 'Log Date:'")
        print("marker to determine their date — they were still imported, tagged with a")
        print("placeholder date (UNKNOWN-ROW-<excel row #>) so nothing was lost. You can")
        print("find and fix these in DB Browser for SQLite by searching date_filled for 'UNKNOWN'.")
    if filled_count:
        print(f"Filled in blank manual fields on {filled_count} existing row(s).")
    print("Backfill complete.")


if __name__ == "__main__":
    main()
